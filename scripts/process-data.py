"""
CX Imperatives 2026 – Survey Data Processor
Converts the Excel survey data tables into a structured JSON file
for use by the Data Explorer web application.

Usage:
    python scripts/process-data.py

Input:  scripts/data/989-25 CX Imperatives 2026 - Survey Data Tables (Topline).xlsx
        (or the root-level copy of the same file)
Output: app/data/survey-data.json
"""

import json
import re
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(SCRIPT_DIR)

EXCEL_CANDIDATES = [
    os.path.join(SCRIPT_DIR, "data", "989-25 CX Imperatives 2026 - Survey Data Tables (Topline).xlsx"),
    os.path.join(ROOT_DIR, "989-25 CX Imperatives 2026 - Survey Data Tables (Topline).xlsx"),
]

OUTPUT_PATH = os.path.join(ROOT_DIR, "docs", "data", "survey-data.json")

SHEETS_TO_PROCESS = [
    ("Screening & Classification", "screening"),
    ("CX Activities & Expectations", "cx-activities"),
    ("Personalization & High-Touch CX", "personalization"),
    ("Digital Experiences & AI", "digital-ai"),
]

SECTION_TITLES = {
    "screening": "Screening & Classification",
    "cx-activities": "CX Activities & Expectations",
    "personalization": "Personalization & High-Touch CX",
    "digital-ai": "Digital Experiences & AI",
}

BRAND_CATEGORIES = [
    "retailer", "packaged goods brand", "clothing or textile brand",
    "financial services provider", "insurance provider", "healthcare provider",
    "household (durable) goods brand", "outdoor recreation or sporting goods brand",
    "vehicle manufacturer", "electronics (hardware) brand", "software brand",
    "telecommunications service provider", "travel or transportation company",
    "restaurant or foodservice brand", "hospitality brand", "media brand",
    "entertainment brand", "nonprofit organization",
]

BRAND_SECTORS = [
    "RETAIL & CPG", "FINSERV & INSURANCE", "HEALTHCARE",
    "AUTO & MFG", "TECH & TELECOM", "TRAVEL & HOSPITALITY",
    "MEDIA & ENTERTAINMENT", "NONPROFIT",
]

# ---------------------------------------------------------------------------
# Text cleaning helpers
# ---------------------------------------------------------------------------

PIPE_PATTERNS = [
    (r'\[made a purchase from\s*/\s*made a donation to\]', 'purchased from / donated to'),
    (r'\[CX ACTIVITY\]', 'engage'),
    (r'\[ASSIGNED BRAND TYPE\]s?', 'brands'),
    (r'\[Q10 CATEGORY\]', 'your assigned brand category'),
    (r'\[BRAND CATEGORY\]', 'your brand category'),
    (r'\[Q10 BRAND\]', 'brands'),
    (r'\[PIPE Q10 CATEGORY(?:\s+NAME)?\]s?', 'brands'),
    (r'\[Q10\.?\s*GROUP\s*\(Brand Sector Assigned\)\]', ''),
    (r'\s*by\s+Q10\.\s*CATEGORY\s*\(Brand Category Assigned\)', ''),
    (r'\s*by\s+Q10\.\s*GROUP\s*\(Brand Sector Assigned\)', ''),
    (r'\s*by\s+Q8\.\s*CX ACTIVITY\s*Assigned', ''),
    (r'\xa0', ' '),
    (r'  +', ' '),
    (r'[\u2019\u2018]', "'"),
    (r'[\u201C\u201D]', '"'),
    (r'[\u2013\u2014]', '-'),
    (r'\u2026', '...'),
    (r'[^\x00-\x7F\u00C0-\u024F\u2018-\u2026]', ''),  # strip remaining odd chars
]


def clean_text(text: str) -> str:
    if not text:
        return ""
    text = str(text).strip()
    for pattern, replacement in PIPE_PATTERNS:
        text = re.sub(pattern, replacement, text, flags=re.IGNORECASE)
    text = text.strip()
    # Normalise ellipsis characters that appear as ?
    text = re.sub(r'\?+\s*\)', '?)', text)
    text = re.sub(r'\?+$', '?', text)
    return text


def to_float(val) -> float | None:
    if val is None:
        return None
    try:
        f = float(val)
        return round(f, 6)
    except (TypeError, ValueError):
        return None


def is_question_row(cell_val) -> bool:
    """True if the cell starts a new question block, e.g. 'Q12. ...'."""
    if not cell_val:
        return False
    s = str(cell_val).strip()
    return bool(re.match(r'^Q\d+[\.\s]', s))


def extract_question_number(cell_val) -> str:
    """Extract 'Q12' from 'Q12. Please think...'."""
    m = re.match(r'^(Q\d+)', str(cell_val).strip())
    return m.group(1) if m else ""


# ---------------------------------------------------------------------------
# Column classification helpers
# ---------------------------------------------------------------------------

def classify_columns(header_row: list) -> tuple[str, list[str]]:
    """
    Determine the cross-tab type from a header row.
    Returns (type_str, [column_names])
    type_str is one of: 'topline', 'by_category', 'by_sector', 'other'
    """
    if not header_row or header_row[0] is None:
        return "topline", []

    first = str(header_row[0]).strip() if header_row[0] else ""

    # Topline marker: first cell is a space, second is '%'
    if first in (' ', '') and len(header_row) > 1 and str(header_row[1]).strip() == '%':
        return "topline", ["NET"]

    if first == "Column %":
        cols = [str(c).strip() for c in header_row[1:] if c is not None]
        # Detect by_sector: all known sector names present
        sector_hits = sum(1 for c in cols if c.upper() in [s.upper() for s in BRAND_SECTORS])
        cat_hits = sum(1 for c in cols if c.lower() in [b.lower() for b in BRAND_CATEGORIES])
        if sector_hits >= 4:
            return "by_sector", cols
        if cat_hits >= 4:
            return "by_category", cols
        # fallback – return as generic cross-tab
        return "by_category", cols

    return "topline", []


# ---------------------------------------------------------------------------
# Core parser
# ---------------------------------------------------------------------------

def parse_sheet(ws) -> list[dict]:
    """
    Parse a worksheet and return a list of raw question dicts.
    """
    rows = [row for row in ws.iter_rows(values_only=True)]
    questions: list[dict] = []
    i = 0
    n = len(rows)

    while i < n:
        row = rows[i]
        cell0 = row[0] if row[0] is not None else ""

        if not is_question_row(cell0):
            i += 1
            continue

        q_num = extract_question_number(cell0)
        q_text_raw = str(cell0).strip()
        # Remove the 'Qnn. ' prefix from the text for cleanliness
        q_text = re.sub(r'^Q\d+\.\s*', '', q_text_raw).strip()
        q_text = clean_text(q_text)

        # Look ahead for the header row: it must be a 'Column %' or topline marker.
        # There may be 1-2 intermediate subheader rows (e.g. 'NET - DIGITAL [...]')
        # that we must skip.
        header_row = []
        tab_type, col_names = "topline", []
        for look_ahead in range(1, 5):
            if i + look_ahead >= n:
                break
            candidate = list(rows[i + look_ahead])
            t, c = classify_columns(candidate)
            if t in ("by_category", "by_sector"):
                header_row = candidate
                tab_type, col_names = t, c
                i += look_ahead
                break
            first_cell = str(candidate[0]).strip() if candidate[0] else ""
            second_cell = str(candidate[1]).strip() if len(candidate) > 1 and candidate[1] else ""
            if first_cell in (' ', '') and second_cell == '%':
                header_row = candidate
                tab_type, col_names = "topline", ["NET"]
                i += look_ahead
                break
        else:
            i += 1  # Advance at least one row if no header found

        # Collect response rows until next question / footer / blank section
        i += 1
        responses: list[dict] = []
        net_values: dict | None = None
        col_n: dict = {}

        while i < n:
            row = rows[i]
            label_raw = row[0]

            # Stop conditions
            if label_raw is None:
                # Blank row – skip but don't stop yet (there may be gaps)
                i += 1
                blank_count = 0
                while i < n and rows[i][0] is None:
                    blank_count += 1
                    i += 1
                    if blank_count > 2:
                        break
                continue

            label_str = str(label_raw).strip()

            if label_str in ('Back to TOC',):
                i += 1
                break

            if is_question_row(label_str):
                break  # don't increment – outer loop will handle

            if label_str.startswith('Total sample') or label_str.startswith('Multiple comparison') or label_str.startswith('Filter:'):
                i += 1
                continue

            if label_str == 'NET':
                if tab_type == 'topline':
                    # The NET value is in col index 1
                    net_val = to_float(row[1]) if len(row) > 1 else None
                    net_values = {"NET": net_val}
                else:
                    net_values = {}
                    for ci, col in enumerate(col_names):
                        val = to_float(row[ci + 1]) if ci + 1 < len(row) else None
                        if val is not None:
                            net_values[col] = val
                i += 1
                continue

            if label_str == 'Column n':
                for ci, col in enumerate(col_names):
                    try:
                        v = row[ci + 1]
                        col_n[col] = int(v) if v is not None else None
                    except (TypeError, ValueError):
                        col_n[col] = None
                i += 1
                continue

            # Regular response row
            label = clean_text(label_str)
            if not label:
                i += 1
                continue

            if tab_type == 'topline':
                val = to_float(row[1]) if len(row) > 1 else None
                responses.append({"label": label, "values": {"NET": val}})
            else:
                values: dict = {}
                for ci, col in enumerate(col_names):
                    val = to_float(row[ci + 1]) if ci + 1 < len(row) else None
                    values[col] = val
                responses.append({"label": label, "values": values})

            i += 1

        if responses:
            questions.append({
                "questionNumber": q_num,
                "text": q_text,
                "rawText": q_text_raw,
                "type": tab_type,
                "crossTabColumns": col_names if tab_type != "topline" else None,
                "responses": responses,
                "netValues": net_values,
                "columnN": col_n if col_n else None,
            })

    return questions


# ---------------------------------------------------------------------------
# Post-processing: group sub-questions
# ---------------------------------------------------------------------------

Q13_ACTIVITIES = [
    "SHOP (in store or online)",
    "Shop at a store",
    "Shop on a website or mobile app",
    "Make a purchase (either in-person or online)",
    "Use / consume a product or service",
    "Start / renew a subscription or service agreement",
    "Join or participate in a rewards program",
    "Contact customer service or request technical support",
    "Donated goods or money",
    "[CX ACTIVITY]",
]


def extract_activity_from_text(text: str) -> str | None:
    """Extract the CX activity phrase from a Q13/Q14/Q15 question text."""
    # Patterns: "when you SHOP (in store or online) with..."
    #           "when you Shop at a store with..."
    #           "Thinking about the last time you SHOPPED (in store..."
    m = re.search(
        r'(?:when you|time you)\s+(.*?)\s+(?:with|[A-Z]{2,})',
        text, re.IGNORECASE
    )
    if m:
        return m.group(1).strip()
    # Fallback: extract text inside brackets in the raw question
    m2 = re.search(r'\[TOP 2 BOX[^\]]*\].*?you\s+(.*?)\s+with', text, re.IGNORECASE)
    if m2:
        return m2.group(1).strip()
    return None


def extract_activity_label(q_text_raw: str) -> str:
    """
    Extract a clean activity label from the raw question text.
    E.g. 'Q13. [TOP 2 BOX...] In general, when you SHOP (in store or online) with...'
    -> 'SHOP (in store or online)'
    """
    # Remove the Q-number prefix and bracket annotations first
    text = re.sub(r'^Q\d+\.\s*', '', q_text_raw).strip()
    text = re.sub(r'\[TOP\s+2\s+BOX[^\]]*\]', '', text, flags=re.IGNORECASE).strip()

    # Pattern 1: "when you <ACTIVITY> with/from [BRAND]"
    # Stops at "with" or "from" followed by a bracket or known brand phrase
    m = re.search(
        r'when you\s+(?:do\s+)?(.*?)\s+(?:with|from)\s+\[(?:BRAND|Q10)',
        text, re.IGNORECASE
    )
    if m:
        act = m.group(1).strip()
        act = re.sub(r'\[.*?\]', '', act).strip()
        if act and act.lower() not in ('[cx activity]', ''):
            return act
        return "Overall (All Activities)"

    # Pattern 2: "when you <ACTIVITY> with/from brands" (after pipe substitution)
    m2 = re.search(r'when you\s+(?:do\s+)?(.*?)\s+(?:with|from)\s+\w', text, re.IGNORECASE)
    if m2:
        act = m2.group(1).strip()
        act = re.sub(r'\[.*?\]', '', act).strip()
        if act and len(act) > 2:
            return act

    # Pattern 3: "last time you <ACTIVITY> with"
    m3 = re.search(r'last time you\s+((?:did\s+)?)(.*?)\s+(?:with|how)', text, re.IGNORECASE)
    if m3:
        act = m3.group(2).strip()
        act = re.sub(r'\[.*?\]', '', act).strip()
        if act and len(act) > 2:
            return act

    return "Overall (All Activities)"


def group_sub_questions(questions: list[dict]) -> list[dict]:
    """
    Group repeated Q13/Q14/Q15 blocks (each with different CX activity)
    into a single question with sub-questions keyed by activity.
    """
    grouped: dict[str, dict] = {}
    result: list[dict] = []

    for q in questions:
        q_num = q["questionNumber"]

        if q_num not in ("Q13", "Q14", "Q15"):
            result.append(q)
            continue

        activity = extract_activity_label(q["rawText"])

        if q_num not in grouped:
            # Create the parent question
            parent_text = {
                "Q13": "How important is it for the experience to be... (Top 2 Box: Very/Extremely Important)",
                "Q14": "Agreement with statements about the last brand experience (Top 2 Box: Somewhat/Strongly Agree)",
                "Q15": "Post-experience outcomes after last brand interaction (Top 2 Box: Somewhat/Strongly Agree)",
            }.get(q_num, q["text"])

            grouped[q_num] = {
                "questionNumber": q_num,
                "text": parent_text,
                "type": "by_sector_multi_activity",
                "crossTabColumns": q.get("crossTabColumns"),
                "activities": {},
            }

        act_key = activity.strip()
        if act_key not in grouped[q_num]["activities"]:
            grouped[q_num]["activities"][act_key] = {
                "activity": act_key,
                "responses": q["responses"],
                "netValues": q.get("netValues"),
                "columnN": q.get("columnN"),
            }

    # Insert grouped questions in order at the position of first occurrence
    seen_grouped: set[str] = set()
    final: list[dict] = []
    for q in questions:
        q_num = q["questionNumber"]
        if q_num in grouped:
            if q_num not in seen_grouped:
                seen_grouped.add(q_num)
                g = grouped[q_num]
                g["activities"] = list(g["activities"].values())
                final.append(g)
        else:
            final.append(q)

    return final


# ---------------------------------------------------------------------------
# Duplicate Q17/Q18/Q19/Q25/Q26/Q30/Q31/Q32/Q33 (by_category + by_sector)
# ---------------------------------------------------------------------------

def merge_dual_crosstab(questions: list[dict]) -> list[dict]:
    """
    Some questions appear twice: once by_category, once by_sector.
    Merge them into a single question with both dimension sets.
    """
    merged: dict[str, dict] = {}
    result: list[dict] = []
    seen: set[str] = set()

    for q in questions:
        q_num = q["questionNumber"]
        q_type = q.get("type", "topline")

        # These question numbers appear twice with different cross-tab types
        DUAL_XTAB_QS = {"Q17", "Q18", "Q19", "Q25", "Q26", "Q30", "Q31", "Q32", "Q33"}

        if q_num in DUAL_XTAB_QS and q_type in ("by_category", "by_sector"):
            if q_num not in merged:
                merged[q_num] = {
                    "questionNumber": q_num,
                    "text": q["text"],
                    "type": "dual_crosstab",
                    "byCategory": None,
                    "bySector": None,
                }

            if q_type == "by_category":
                merged[q_num]["byCategory"] = {
                    "crossTabColumns": q.get("crossTabColumns"),
                    "responses": q["responses"],
                    "netValues": q.get("netValues"),
                    "columnN": q.get("columnN"),
                }
            elif q_type == "by_sector":
                merged[q_num]["bySector"] = {
                    "crossTabColumns": q.get("crossTabColumns"),
                    "responses": q["responses"],
                    "netValues": q.get("netValues"),
                    "columnN": q.get("columnN"),
                }
        else:
            result.append(q)

    # Insert merged questions at position of first occurrence
    seen_merged: set[str] = set()
    final: list[dict] = []
    for q in questions:
        q_num = q["questionNumber"]
        if q_num in merged:
            if q_num not in seen_merged:
                seen_merged.add(q_num)
                final.append(merged[q_num])
        else:
            final.append(q)

    return final


# ---------------------------------------------------------------------------
# Build clean IDs
# ---------------------------------------------------------------------------

def build_id(q_num: str, existing: set[str]) -> str:
    base = q_num.lower().replace(".", "-")
    if base not in existing:
        existing.add(base)
        return base
    idx = 2
    while f"{base}-{idx}" in existing:
        idx += 1
    new_id = f"{base}-{idx}"
    existing.add(new_id)
    return new_id


# ---------------------------------------------------------------------------
# Main build
# ---------------------------------------------------------------------------

def build_json(wb) -> dict:
    sections = []
    all_ids: set[str] = set()

    for sheet_name, section_id in SHEETS_TO_PROCESS:
        if sheet_name not in wb.sheetnames:
            print(f"  Warning: sheet '{sheet_name}' not found – skipping.")
            continue

        ws = wb[sheet_name]
        print(f"  Processing '{sheet_name}'...")
        raw_questions = parse_sheet(ws)
        print(f"    Found {len(raw_questions)} raw question blocks")

        # Group sub-questions (Q13/Q14/Q15)
        questions = group_sub_questions(raw_questions)
        # Merge dual cross-tab variants
        questions = merge_dual_crosstab(questions)
        print(f"    After grouping: {len(questions)} questions")

        # Assign clean IDs and clean up the dicts
        clean_questions = []
        for q in questions:
            q_id = build_id(q["questionNumber"], all_ids)
            clean_q = {k: v for k, v in q.items() if k not in ("rawText",)}
            clean_q["id"] = q_id
            clean_questions.append(clean_q)

        sections.append({
            "id": section_id,
            "title": SECTION_TITLES[section_id],
            "questions": clean_questions,
        })

    return {
        "metadata": {
            "title": "CX Imperatives 2026",
            "subtitle": "Consumer Survey – Interactive Data Explorer",
            "description": "Merkle's annual consumer experience research covering 2,500 respondents across 30 countries on CX expectations, personalization, and AI.",
            "totalRespondents": 2500,
            "fieldYear": "2025",
            "publisher": "Merkle / dentsu",
        },
        "sections": sections,
        "dimensions": {
            "categories": BRAND_CATEGORIES,
            "sectors": BRAND_SECTORS,
        },
    }


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    # Find the Excel file
    excel_path = None
    for candidate in EXCEL_CANDIDATES:
        if os.path.exists(candidate):
            excel_path = candidate
            break

    if not excel_path:
        print(f"ERROR: Could not find the Excel source file.")
        print(f"Looked in:\n" + "\n".join(f"  {c}" for c in EXCEL_CANDIDATES))
        sys.exit(1)

    print(f"Reading: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    print("Building JSON...")
    data = build_json(wb)

    # Write output
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"\nOutput written to: {OUTPUT_PATH}")

    # Stats
    total_q = sum(len(s["questions"]) for s in data["sections"])
    print(f"Total questions: {total_q}")
    for s in data["sections"]:
        print(f"  {s['title']}: {len(s['questions'])} questions")


if __name__ == "__main__":
    main()
